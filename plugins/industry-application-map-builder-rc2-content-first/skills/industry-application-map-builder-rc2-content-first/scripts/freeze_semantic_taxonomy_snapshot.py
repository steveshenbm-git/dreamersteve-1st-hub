#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import sys

from openpyxl import load_workbook


REQUIRED_HEADERS = {"taxonomy_node_id", "parent_node_id", "status"}


def fail(code: str, detail: str) -> int:
    print(json.dumps({"status": "FAIL", "code": code, "detail": detail}, ensure_ascii=False), file=sys.stderr)
    return 1


def canonical(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def main() -> int:
    parser = argparse.ArgumentParser(description="Freeze current leaf nodes from the industry taxonomy workbook.")
    parser.add_argument("--taxonomy-workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    source = args.taxonomy_workbook.resolve()
    output = args.output.resolve()
    if not source.is_file():
        return fail("TAXONOMY_WORKBOOK_MISSING", str(source))
    if output.exists():
        return fail("OUTPUT_EXISTS", str(output))
    workbook = load_workbook(source, read_only=True, data_only=True)
    if "行业骨架" not in workbook.sheetnames:
        return fail("TAXONOMY_SHEET_MISSING", "行业骨架")
    sheet = workbook["行业骨架"]
    headers = [cell.value for cell in sheet[1]]
    header_set = {value for value in headers if value}
    missing = sorted(REQUIRED_HEADERS - header_set)
    if missing:
        return fail("TAXONOMY_HEADERS_MISSING", ",".join(missing))

    nodes: list[dict] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        record = {header: value for header, value in zip(headers, row) if header}
        node_id = record.get("taxonomy_node_id")
        if not node_id:
            continue
        status = str(record.get("status") or "").strip().lower()
        if status not in {"current", "active"}:
            continue
        nodes.append({key: record.get(key) for key in headers if key})
    if not nodes:
        return fail("NO_CURRENT_TAXONOMY_NODES", str(source))

    parent_ids = {str(node.get("parent_node_id")) for node in nodes if node.get("parent_node_id")}
    terminal_nodes = sorted(
        (node for node in nodes if str(node["taxonomy_node_id"]) not in parent_ids),
        key=lambda item: str(item["taxonomy_node_id"]),
    )
    terminal_hash = hashlib.sha256(canonical(terminal_nodes).encode("utf-8")).hexdigest()
    payload = {
        "schema_version": "1.0",
        "taxonomy_workbook_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "terminal_node_count": len(terminal_nodes),
        "terminal_nodes_sha256": terminal_hash,
        "terminal_nodes": terminal_nodes,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": "PASS", "terminal_node_count": len(terminal_nodes), "terminal_nodes_sha256": terminal_hash}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
