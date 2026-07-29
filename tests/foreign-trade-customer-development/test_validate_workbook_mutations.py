from hashlib import sha256
from pathlib import Path
from shutil import copyfile
import subprocess
import sys
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = (
    REPO_ROOT
    / "plugins"
    / "foreign-trade-customer-development"
    / "skills"
    / "foreign-trade-customer-development"
    / "assets"
    / "prospect-development-workbook.xlsx"
)
VALIDATOR_PATH = Path(__file__).with_name("validate_workbook.py")
HANDOFF_TARGETS = (
    ("客户总览", "R3:R5000"),
    ("移交记录", "J3:J5000"),
)
EXPECTED_HANDOFF_FORMULA = '"未触发,待客户经营与沟通,已移交,业务员已决定"'
OBSOLETE_HANDOFF_FORMULA = '"未触发,触达已暂停,待客户经营与沟通,已移交,业务员已决定"'


def file_sha256(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def validation_by_range(sheet, expected_range: str):
    matches = [
        validation
        for validation in sheet.data_validations.dataValidation
        if str(validation.sqref) == expected_range
    ]
    if len(matches) != 1:
        raise AssertionError(
            f"expected exactly one validation for {sheet.title}.{expected_range}, "
            f"observed {len(matches)}"
        )
    return matches[0]


def run_validator(path: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(VALIDATOR_PATH), str(path)],
        check=False,
        capture_output=True,
        text=True,
    )


def assert_rejected(result: subprocess.CompletedProcess[str], diagnostic: str) -> None:
    if result.returncode == 0:
        raise AssertionError(
            f"validator false-passed mutation {diagnostic!r}:\n{result.stdout}{result.stderr}"
        )
    combined = result.stdout + result.stderr
    diagnostics = [
        line
        for line in combined.splitlines()
        if line.startswith("FAIL ") and not line.startswith("FAIL:")
    ]
    if len(diagnostics) != 1 or not diagnostics[0].startswith(f"FAIL {diagnostic}"):
        raise AssertionError(
            f"validator did not isolate mutation diagnostic {diagnostic!r}:\n{combined}"
        )


def assert_accepted(result: subprocess.CompletedProcess[str], label: str) -> None:
    if result.returncode != 0 or not result.stdout.startswith("PASS:"):
        raise AssertionError(
            f"validator rejected GREEN control {label!r}:\n{result.stdout}{result.stderr}"
        )


def create_green_control(destination: Path) -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    validations = [
        validation
        for sheet in workbook.worksheets
        for validation in sheet.data_validations.dataValidation
    ]
    if len(validations) != 21:
        raise AssertionError(
            f"expected 21 controlled validations in source workbook, observed {len(validations)}"
        )
    for validation in validations:
        validation.showErrorMessage = True
        validation.errorStyle = "stop"
        validation.allowBlank = True
    for sheet_name, expected_range in HANDOFF_TARGETS:
        validation = validation_by_range(workbook[sheet_name], expected_range)
        validation.formula1 = EXPECTED_HANDOFF_FORMULA
    workbook.save(destination)


def main() -> int:
    source_hash = file_sha256(WORKBOOK_PATH)

    with TemporaryDirectory(prefix="ft-workbook-mutations-") as temp_dir:
        temp_root = Path(temp_dir)

        green_control_path = temp_root / "green-control.xlsx"
        create_green_control(green_control_path)
        assert_accepted(run_validator(green_control_path), "all 21 enforced validations")

        route_decision_green_path = temp_root / "route-decision-green-control.xlsx"
        copyfile(green_control_path, route_decision_green_path)
        assert_accepted(
            run_validator(route_decision_green_path),
            "route-decision mutation baseline",
        )
        wrong_route_decision_path = temp_root / "wrong-route-decision-list.xlsx"
        workbook = load_workbook(route_decision_green_path, data_only=False)
        validation = validation_by_range(workbook["路线评审"], "Q3:Q5000")
        validation.formula1 = '"选择编译,AI自动选择,暂缓,淘汰"'
        workbook.save(wrong_route_decision_path)
        assert_rejected(
            run_validator(wrong_route_decision_path),
            "路线评审.salesperson_route_decision",
        )

        screening_green_control_path = temp_root / "screening-green-control.xlsx"
        copyfile(green_control_path, screening_green_control_path)
        assert_accepted(
            run_validator(screening_green_control_path),
            "screening mutation baseline",
        )
        wrong_screening_path = temp_root / "wrong-screening-list.xlsx"
        workbook = load_workbook(screening_green_control_path, data_only=False)
        validation = validation_by_range(workbook["客户总览"], "H3:H5000")
        validation.formula1 = '"待业务员筛选,已确认,NOT_A_VALID_SCREENING_STATE,已关闭"'
        workbook.save(wrong_screening_path)
        assert_rejected(
            run_validator(wrong_screening_path),
            "客户总览.screening_status",
        )

        risk_green_control_path = temp_root / "risk-green-control.xlsx"
        copyfile(green_control_path, risk_green_control_path)
        assert_accepted(
            run_validator(risk_green_control_path),
            "risk-alert mutation baseline",
        )
        disabled_risk_alert_path = temp_root / "disabled-risk-alert.xlsx"
        workbook = load_workbook(risk_green_control_path, data_only=False)
        validation = validation_by_range(workbook["客户总览"], "K3:K5000")
        validation.showErrorMessage = False
        workbook.save(disabled_risk_alert_path)
        assert_rejected(
            run_validator(disabled_risk_alert_path),
            "客户总览.risk_gate: expected showErrorMessage=True",
        )

        blank_green_control_path = temp_root / "blank-green-control.xlsx"
        copyfile(green_control_path, blank_green_control_path)
        assert_accepted(
            run_validator(blank_green_control_path),
            "blank-allowance mutation baseline",
        )
        disabled_blank_allowance_path = temp_root / "disabled-blank-allowance.xlsx"
        workbook = load_workbook(blank_green_control_path, data_only=False)
        validation = validation_by_range(workbook["客户总览"], "H3:H5000")
        validation.allowBlank = False
        workbook.save(disabled_blank_allowance_path)
        assert_rejected(
            run_validator(disabled_blank_allowance_path),
            "客户总览.screening_status: expected allowBlank=True",
        )

        for index, (sheet_name, expected_range) in enumerate(HANDOFF_TARGETS, start=1):
            handoff_green_control_path = temp_root / f"handoff-green-control-{index}.xlsx"
            copyfile(green_control_path, handoff_green_control_path)
            assert_accepted(
                run_validator(handoff_green_control_path),
                f"{sheet_name}.handoff_status mutation baseline",
            )
            obsolete_handoff_state_path = temp_root / f"obsolete-handoff-state-{index}.xlsx"
            workbook = load_workbook(handoff_green_control_path, data_only=False)
            validation = validation_by_range(workbook[sheet_name], expected_range)
            validation.formula1 = OBSOLETE_HANDOFF_FORMULA
            workbook.save(obsolete_handoff_state_path)
            assert_rejected(
                run_validator(obsolete_handoff_state_path),
                f"{sheet_name}.handoff_status",
            )

    if file_sha256(WORKBOOK_PATH) != source_hash:
        raise AssertionError("source workbook changed during mutation tests")

    print("PASS: GREEN control validates and all six isolated mutations are rejected")
    return 0


class WorkbookMutationTests(unittest.TestCase):
    def test_validator_rejects_isolated_workbook_contract_mutations(self) -> None:
        self.assertEqual(main(), 0)


if __name__ == "__main__":
    raise SystemExit(main())
