from pathlib import Path
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

EXPECTED = {
    "路线评审": ["route_review_id", "source_route_candidate_id", "company_id", "product_scope", "route_packet_reference", "route_packet_sha256", "producer_registry_reference", "map_route_status", "research_readiness", "readiness_snapshot_reference", "knowledge_snapshot_hash", "readiness_fact_ids", "commercial_readiness_status", "readiness_reviewed_at", "stale_status", "unresolved_business_conditions", "salesperson_route_decision", "decision_basis", "decision_date"],
    "开发方向": ["direction_id", "source_route_review_id", "source_route_candidate_id", "approved_product_reference", "product_boundary", "observable_enterprise_rule", "candidate_direct_evidence_rule", "exclusion_boundary", "external_validation_status", "direction_status", "declared_scope", "unresolved_conditions", "salesperson_decision", "decision_date"],
    "方向证据": ["direction_evidence_id", "direction_id", "source_type", "source_title", "source_url_or_local_reference", "original_excerpt", "zh_summary", "published_at", "observed_at", "evidence_state", "validation_effect", "limitation_note"],
    "客户总览": ["customer_id", "company_name", "legal_name", "website", "country", "business_model", "source_direction_id", "screening_status", "salesperson_classification", "information_reliability", "risk_gate", "recommended_opportunity_id", "primary_contact_id", "last_research_date", "last_touch_date", "next_action", "next_action_date", "handoff_status", "salesperson_notes"],
    "公司研究": ["research_id", "customer_id", "research_level", "section", "finding_original", "finding_zh_summary", "evidence_id", "evidence_state", "source_published_at", "observed_at", "gap_or_conflict", "salesperson_confirmed"],
    "项目机会": ["opportunity_id", "customer_id", "customer_fact", "application_or_purchase_scenario", "approved_product_reference", "fit_basis", "validation_question", "primary_contact_id", "recommendation_state", "salesperson_decision", "decision_date"],
    "联系人": ["contact_id", "customer_id", "name", "title", "possible_role", "role_evidence_id", "channel", "contact_value", "authenticity_state", "source_reliability", "usage_permission", "contact_order", "ordering_basis", "observed_at", "salesperson_approval", "employer_or_entity", "entity_match_basis", "contact_source_reference", "uncertainty_note"],
    "证据来源": ["evidence_id", "customer_id", "source_type", "source_title", "source_url_or_local_reference", "source_owner", "source_language", "original_excerpt", "zh_summary", "published_at", "observed_at", "evidence_state", "access_scope", "conflict_note", "source_region_or_jurisdiction"],
    "海关与贸易": ["trade_record_id", "customer_id", "data_source", "access_scope", "coverage_country", "coverage_period", "observed_entity_name", "observed_entity_address", "entity_match_basis", "trade_direction", "shipment_period", "visible_frequency", "product_description", "hs_code", "quantity", "quantity_unit", "weight", "weight_unit", "declared_value", "declared_currency", "partner_or_country", "limitation_note", "observed_at", "evidence_id"],
    "风险核验": ["risk_id", "customer_id", "risk_type", "matched_entity", "match_basis", "allegation_or_record", "evidence_id", "jurisdiction", "record_date", "observed_at", "evidence_state", "false_match_risk", "gate_status", "reviewer_decision", "decision_date"],
    "触达记录": ["touch_id", "customer_id", "contact_id", "channel", "touch_stage", "content_status", "planned_date", "draft_content_or_local_reference", "draft_generated_at", "draft_for_touch_stage", "automation_run_id", "actual_sent_at", "actual_content_or_local_reference", "response_state", "response_at", "next_action", "next_action_date", "salesperson_approved"],
    "移交记录": ["handoff_id", "customer_id", "trigger_channel", "trigger_touch_id", "response_reference", "development_snapshot_reference", "open_questions", "risk_gate_status", "target_skill", "handoff_status", "salesperson_decision", "decision_date"],
}

EXPECTED_ZH = {
    "路线评审": ["路线评审编号", "来源路线候选编号", "公司编号", "产品范围", "路线包引用", "路线包哈希", "生产者登记引用", "地图路线状态", "研究就绪状态", "承接视图引用", "知识快照哈希", "承接事实编号", "商业承接状态", "承接复核日期", "时效状态", "未解决业务条件", "业务员路线决定", "决定依据", "决定日期"],
    "开发方向": ["开发方向编号", "来源路线评审编号", "来源路线候选编号", "已批准产品引用", "产品边界", "可观察目标企业规则", "候选公司直接证据规则", "排除边界", "外部核实状态", "方向状态", "本次声明范围", "待核实条件", "业务员方向决定", "决定日期"],
    "方向证据": ["方向证据编号", "开发方向编号", "来源类型", "来源标题", "来源网址或本地引用", "原文摘录", "中文摘要", "来源发布日期", "观察记录时间", "证据状态", "对方向的验证作用", "局限说明"],
    "客户总览": ["客户编号", "公司常用名称", "法定注册名称", "公司网站", "国家或地区", "商业模式", "来源开发方向编号", "筛选状态", "业务员客户分类", "信息可靠性", "风险门状态", "推荐机会编号", "主要联系人编号", "最近研究日期", "最近触达日期", "下一步行动", "下一步行动日期", "移交状态", "业务员备注"],
    "公司研究": ["研究记录编号", "客户编号", "研究层级", "研究章节", "原文研究发现", "研究发现中文摘要", "证据编号", "证据状态", "来源发布日期", "观察记录时间", "信息缺口或冲突", "业务员确认状态"],
    "项目机会": ["机会编号", "客户编号", "已确认客户事实", "应用或采购场景", "已批准产品引用", "匹配依据", "待验证问题", "主要联系人编号", "推荐状态", "业务员决定", "决定日期"],
    "联系人": ["联系人编号", "客户编号", "联系人姓名", "职务名称", "可能承担的角色", "角色证据编号", "联系渠道", "联系方式内容", "真实性状态", "来源可靠性", "使用许可", "联系顺序", "排序依据", "观察记录时间", "业务员批准状态", "所属公司或主体", "主体匹配依据", "联系信息来源或职业页面", "身份职位或联系方式不确定项"],
    "证据来源": ["证据编号", "客户编号", "来源类型", "来源标题", "来源网址或本地引用", "来源主体", "来源语言", "原文摘录", "中文摘要", "来源发布日期", "观察记录时间", "证据状态", "访问范围", "冲突说明", "来源适用地区或管辖范围"],
    "海关与贸易": ["贸易记录编号", "客户编号", "数据来源", "访问范围", "覆盖国家或地区", "覆盖期间", "观察到的企业名称", "观察到的企业地址", "主体匹配依据", "贸易方向", "货运期间", "可见交易频次", "产品描述", "海关编码", "数量", "数量单位", "重量", "重量单位", "申报价值", "申报币种", "贸易伙伴或国家", "数据局限说明", "观察记录时间", "证据编号"],
    "风险核验": ["风险记录编号", "客户编号", "风险类型", "匹配到的主体", "主体匹配依据", "指控或记录内容", "证据编号", "管辖地区", "记录日期", "观察记录时间", "证据状态", "误匹配风险", "风险门状态", "审核人决定", "决定日期"],
    "触达记录": ["触达记录编号", "客户编号", "联系人编号", "触达渠道", "触达阶段", "内容状态", "计划日期", "草稿内容或本地引用", "草稿生成时间", "草稿对应触达阶段", "自动化运行编号", "实际发送时间", "实发内容或本地引用", "回复状态", "回复时间", "下一步行动", "下一步行动日期", "业务员批准状态"],
    "移交记录": ["移交记录编号", "客户编号", "触发渠道", "触发触达记录编号", "回复内容引用", "客户开发快照引用", "未解决问题", "风险门状态", "目标技能", "移交状态", "业务员决定", "决定日期"],
}

SCREENING_STATES = ["待业务员筛选", "已确认", "已暂停", "已关闭"]
DIRECTION_STATES = ["草案", "待外部核实", "待业务员确认", "已确认可扫描", "暂缓", "淘汰"]
EXTERNAL_VALIDATION_STATES = ["支持", "存在反证", "证据有限", "尚未核实", "来源不可访问"]
CLASSIFICATION_STATES = ["不继续", "普通候选", "潜力客户"]
RELIABILITY_STATES = [
    "资料充分且一致",
    "整体可信但存在缺口",
    "存在重大冲突需要核验",
    "证据不足无法判断",
]
EVIDENCE_STATES = [
    "官方直接证据",
    "多来源相互印证",
    "单一来源待验证",
    "合理推断",
    "来源相互冲突",
    "信息已经过期",
    "来源不明隔离待核实",
]
USAGE_PERMISSION_STATES = ["正常使用", "限制使用", "隔离待核实"]
RISK_STATES = ["未触发", "待核验", "暂停待业务员审核", "业务员批准继续", "已关闭"]
CONTENT_STATES = ["草稿", "业务员批准", "计划触达", "实际发送", "实际回复"]
HANDOFF_STATES = ["未触发", "待客户经营与沟通", "已移交", "业务员已决定"]
MAP_ROUTE_STATES = ["路线线索", "路线候选", "待外部核实", "暂缓", "排除"]
RESEARCH_READINESS_STATES = ["可编译方向", "需补路线证据", "待外部核实", "不可进入"]
COMMERCIAL_READINESS_STATES = ["可承接", "有条件", "未知", "已确认冲突"]
STALE_STATES = ["当前", "临近复核", "已过期", "无法判断"]
ROUTE_DECISION_STATES = ["选择编译", "继续核实", "暂缓", "淘汰"]

CONTROLLED_VALIDATIONS = {
    ("路线评审", "map_route_status"): ("H3:H5000", MAP_ROUTE_STATES),
    ("路线评审", "research_readiness"): ("I3:I5000", RESEARCH_READINESS_STATES),
    ("路线评审", "commercial_readiness_status"): ("M3:M5000", COMMERCIAL_READINESS_STATES),
    ("路线评审", "stale_status"): ("O3:O5000", STALE_STATES),
    ("路线评审", "salesperson_route_decision"): ("Q3:Q5000", ROUTE_DECISION_STATES),
    ("开发方向", "external_validation_status"): ("I3:I5000", EXTERNAL_VALIDATION_STATES),
    ("开发方向", "direction_status"): ("J3:J5000", DIRECTION_STATES),
    ("方向证据", "evidence_state"): ("J3:J5000", EVIDENCE_STATES),
    ("客户总览", "screening_status"): ("H3:H5000", SCREENING_STATES),
    ("客户总览", "salesperson_classification"): ("I3:I5000", CLASSIFICATION_STATES),
    ("客户总览", "information_reliability"): ("J3:J5000", RELIABILITY_STATES),
    ("客户总览", "risk_gate"): ("K3:K5000", RISK_STATES),
    ("客户总览", "handoff_status"): ("R3:R5000", HANDOFF_STATES),
    ("公司研究", "evidence_state"): ("H3:H5000", EVIDENCE_STATES),
    ("联系人", "usage_permission"): ("K3:K5000", USAGE_PERMISSION_STATES),
    ("证据来源", "evidence_state"): ("L3:L5000", EVIDENCE_STATES),
    ("风险核验", "evidence_state"): ("K3:K5000", EVIDENCE_STATES),
    ("风险核验", "gate_status"): ("M3:M5000", RISK_STATES),
    ("触达记录", "content_status"): ("F3:F5000", CONTENT_STATES),
    ("移交记录", "risk_gate_status"): ("H3:H5000", RISK_STATES),
    ("移交记录", "handoff_status"): ("J3:J5000", HANDOFF_STATES),
}


def list_values(formula):
    if not formula:
        return []
    value = formula.strip()
    if value.startswith('"') and value.endswith('"'):
        value = value[1:-1]
    return [item.strip() for item in value.split(",")]


def validation_ranges(validation):
    for cell_range in str(validation.sqref).split():
        yield range_boundaries(cell_range)


def validation_for_cell(sheet, column_index, row_index):
    matches = []
    for validation in sheet.data_validations.dataValidation:
        for min_col, min_row, max_col, max_row in validation_ranges(validation):
            if min_col <= column_index <= max_col and min_row <= row_index <= max_row:
                matches.append(validation)
                break
    return matches


path = Path(sys.argv[1])
workbook = load_workbook(path, data_only=False)
failures = []

if workbook.sheetnames != list(EXPECTED):
    failures.append(
        f"sheet_order: expected {list(EXPECTED)!r}, observed {workbook.sheetnames!r}"
    )

for name, expected_headers in EXPECTED.items():
    if name not in workbook.sheetnames:
        failures.append(f"{name}: missing worksheet")
        continue
    sheet = workbook[name]
    headers = [cell.value for cell in sheet[1]]
    zh_headers = [cell.value for cell in sheet[2]]
    if headers != expected_headers:
        failures.append(
            f"{name}.row1_headers: expected {expected_headers!r}, observed {headers!r}"
        )
    if zh_headers != EXPECTED_ZH[name]:
        failures.append(
            f"{name}.row2_headers: expected {EXPECTED_ZH[name]!r}, observed {zh_headers!r}"
        )
    if sheet.freeze_panes != "A3":
        failures.append(f"{name}.freeze_panes: expected 'A3', observed {sheet.freeze_panes!r}")
    expected_filter = f"A2:{get_column_letter(len(expected_headers))}2"
    table_filters = [
        table.ref for table in sheet.tables.values()
        if table.ref == expected_filter
    ]
    if sheet.auto_filter.ref != expected_filter and not table_filters:
        failures.append(
            f"{name}.auto_filter: expected worksheet or table filter {expected_filter!r}, "
            f"observed worksheet={sheet.auto_filter.ref!r}, tables={[table.ref for table in sheet.tables.values()]!r}"
        )
    if sheet.max_row != 2:
        failures.append(
            f"{name}.max_row: expected 2 with no real rows from row 3, observed {sheet.max_row}"
        )
    for validation in sheet.data_validations.dataValidation:
        for _, min_row, _, _ in validation_ranges(validation):
            if min_row != 3:
                failures.append(
                    f"{name}.data_validation: {validation.sqref} must begin exactly at row 3"
                )

for (sheet_name, field_name), (expected_range, expected_values) in CONTROLLED_VALIDATIONS.items():
    if sheet_name not in workbook.sheetnames:
        failures.append(f"{sheet_name}.{field_name}: worksheet is missing")
        continue
    sheet = workbook[sheet_name]
    observed_headers = [cell.value for cell in sheet[1]]
    if field_name not in observed_headers:
        failures.append(f"{sheet_name}.{field_name}: field is missing")
        continue
    column_index = observed_headers.index(field_name) + 1
    validations = validation_for_cell(sheet, column_index, 3)
    if len(validations) != 1:
        failures.append(
            f"{sheet_name}.{field_name}: expected one validation at row 3, "
            f"observed {len(validations)}"
        )
        continue
    validation = validations[0]
    if str(validation.sqref) != expected_range:
        failures.append(
            f"{sheet_name}.{field_name}: expected exact range {expected_range!r}, "
            f"observed {str(validation.sqref)!r}"
        )
    if validation.type != "list":
        failures.append(
            f"{sheet_name}.{field_name}: expected list validation, observed {validation.type!r}"
        )
    observed_states = list_values(validation.formula1)
    if observed_states != expected_values:
        failures.append(
            f"{sheet_name}.{field_name}: expected {expected_values!r}, "
            f"observed {observed_states!r}"
        )
    if validation.showErrorMessage is not True:
        failures.append(
            f"{sheet_name}.{field_name}: expected showErrorMessage=True, "
            f"observed {validation.showErrorMessage!r}"
        )
    if validation.errorStyle != "stop":
        failures.append(
            f"{sheet_name}.{field_name}: expected errorStyle='stop', "
            f"observed {validation.errorStyle!r}"
        )
    if validation.allowBlank is not True:
        failures.append(
            f"{sheet_name}.{field_name}: expected allowBlank=True, "
            f"observed {validation.allowBlank!r}"
        )

observed_validation_count = sum(
    len(sheet.data_validations.dataValidation) for sheet in workbook.worksheets
)
if observed_validation_count != len(CONTROLLED_VALIDATIONS):
    failures.append(
        "controlled_validations.count: expected "
        f"{len(CONTROLLED_VALIDATIONS)}, observed {observed_validation_count}"
    )

if failures:
    for failure in failures:
        print(f"FAIL {failure}")
    print(f"FAIL: {len(failures)} workbook contract diagnostics")
    raise SystemExit(1)

print("PASS: workbook structure, bilingual headers, empty-data boundary, freeze panes, "
      "filters, and all controlled validations with stop-style enforcement and blank allowance")
