from __future__ import annotations

import hashlib
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SKILL_ROOT = (
    ROOT
    / "plugins"
    / "industry-application-map-builder"
    / "skills"
    / "industry-application-map-builder"
)
INIT_SCRIPT = SKILL_ROOT / "scripts" / "init_industry_application_workspace.py"
VALIDATE_SCRIPT = SKILL_ROOT / "scripts" / "validate_industry_application_workspace.py"
EXPORT_SCRIPT = SKILL_ROOT / "scripts" / "export_company_route_pool.py"


def run_script(script: Path, *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(script), *args],
        check=False,
        capture_output=True,
        text=True,
    )


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def append_record(path: Path, sheet_name: str, record: dict) -> None:
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    sheet.append([record.get(header) for header in headers])
    workbook.save(path)


def change_first_record(path: Path, sheet_name: str, field: str, value) -> None:
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index(field) + 1
    sheet.cell(row=3, column=column, value=value)
    workbook.save(path)


class WorkspaceToolTests(unittest.TestCase):
    def test_bundled_workbooks_preserve_operational_contract(self):
        assets = SKILL_ROOT / "assets"
        for workbook_path in sorted(assets.glob("*.xlsx")):
            with self.subTest(workbook=workbook_path.name):
                workbook = load_workbook(workbook_path)
                for sheet in workbook.worksheets:
                    self.assertEqual(sheet.freeze_panes, "A3", sheet.title)
                    self.assertEqual(len(sheet.tables), 1, sheet.title)
                    table = next(iter(sheet.tables.values()))
                    self.assertTrue(table.ref.startswith("A2:"), sheet.title)
                    for validation in sheet.data_validations.dataValidation:
                        self.assertTrue(validation.allow_blank, sheet.title)
                        self.assertEqual(validation.errorStyle, "stop", sheet.title)
                        self.assertTrue(validation.showErrorMessage, sheet.title)

    def initialize_root(self, parent: Path) -> Path:
        destination = parent / "industry-map"
        result = run_script(
            INIT_SCRIPT,
            "--mode",
            "root",
            "--destination",
            str(destination),
            "--taxonomy-system",
            "TEST",
            "--taxonomy-version",
            "2026",
            "--taxonomy-source-url",
            "https://example.com/taxonomy",
            "--declared-scope",
            "full-test",
            "--application-base-version",
            "2026.1",
            "--source-scope",
            "public",
        )
        self.assertEqual(result.returncode, 0, result.stderr + result.stdout)
        return destination

    def create_product_library_for(
        self, parent: Path, company_id: str, product_family: str
    ) -> tuple[Path, Path]:
        library = parent / f"{company_id}-product-library"
        facts_path = library / "02-事实库" / "facts.json"
        packet_path = library / "04-开发交接" / "product-development-fact-packet.json"
        fact_id = f"{company_id}-K-0001"
        source_id = f"{company_id}-S-0001"
        write_json(library / "company.json", {"company_id": company_id})
        write_json(
            facts_path,
            {
                "schema_version": "1.0",
                "company_id": company_id,
                "facts": [
                    {
                        "fact_id": fact_id,
                        "company_id": company_id,
                        "subject_scope": "own_company",
                        "statement_kind": "source_fact",
                        "evidence_level": "E3",
                        "review_status": "approved",
                        "source_id": source_id,
                        "allowed_use": ["internal", "external"],
                    }
                ],
            },
        )
        write_json(
            packet_path,
            {
                "schema_version": "1.1",
                "product_development_fact_packet": {
                    "company_id": company_id,
                    "product_family": product_family,
                    "product_series_or_model": None,
                    "confirmed_form": [],
                    "confirmed_parameters": [],
                    "confirmed_properties": [],
                    "confirmed_mechanisms": [],
                    "confirmed_functions": [fact_id],
                    "confirmed_effects": [],
                    "confirmed_applications": [],
                    "required_conditions": [],
                    "known_limits": [],
                    "unresolved_conditions": [],
                    "approved_references": [source_id],
                    "allowed_use": ["internal_industry_application_mapping"],
                    "prohibited_inference": [],
                    "generated_at": "2026-07-29",
                },
            },
        )
        return library, packet_path

    def create_product_library(self, parent: Path) -> tuple[Path, Path]:
        return self.create_product_library_for(
            parent, "ACME-001", "Synthetic control material"
        )

    def initialize_company(
        self,
        root: Path,
        library: Path,
        packet: Path,
        company_id: str = "ACME-001",
        product_scope: str = "Synthetic control material",
    ) -> Path:
        result = run_script(
            INIT_SCRIPT,
            "--mode",
            "company",
            "--map-root",
            str(root),
            "--company-id",
            company_id,
            "--company-library-root",
            str(library),
            "--product-packet",
            str(packet),
            "--product-scope",
            product_scope,
            "--declared-taxonomy-scope",
            "TEST-2026-1000",
            "--declared-application-scope",
            "APP-001",
            "--allowed-source-scope",
            "public",
        )
        self.assertEqual(result.returncode, 0, result.stderr + result.stdout)
        return root / "04-公司地图" / company_id

    def populate_shared_base(self, root: Path) -> None:
        taxonomy = root / "01-共享行业骨架" / "industry-taxonomy.xlsx"
        application = root / "02-共享应用知识" / "industry-application-base.xlsx"

        append_record(
            taxonomy,
            "证据来源",
            {
                "source_id": "APP-S-001",
                "source_type": "official_classification",
                "title": "Synthetic official classification",
                "publisher": "Synthetic Statistics Authority",
                "source_url_or_local_reference": "https://example.com/taxonomy",
                "published_at": "2026-01-01",
                "observed_at": "2026-07-29",
                "access_scope": "public",
                "notes": "Synthetic fixture only",
            },
        )
        append_record(
            taxonomy,
            "行业骨架",
            {
                "taxonomy_node_id": "TEST-2026-1000",
                "taxonomy_system": "TEST",
                "taxonomy_version": "2026",
                "code": "1000",
                "name_zh": "合成行业",
                "name_en": "Synthetic industry",
                "level": "class",
                "parent_node_id": "",
                "valid_from": "2026-01-01",
                "valid_to": "",
                "status": "current",
                "source_id": "APP-S-001",
            },
        )
        append_record(
            application,
            "证据来源",
            {
                "evidence_id": "APP-E-001",
                "source_type": "official_technical_page",
                "title": "Synthetic application evidence",
                "publisher": "Independent Publisher",
                "source_url_or_local_reference": "https://example.com/application",
                "published_at": "2026-01-01",
                "observed_at": "2026-07-29",
                "source_subject": "general_industry",
                "source_dependency_group": "APP-SOURCE-GROUP-1",
                "original_location": "section 1",
                "zh_summary": "合成应用证据",
                "evidence_state": "supported",
                "access_scope": "public",
                "conflict_note": "",
            },
        )
        append_record(
            application,
            "产出产品",
            {
                "output_product_id": "OUT-001",
                "name_zh": "合成产出产品",
                "name_en": "Synthetic output",
                "aliases": "[]",
                "taxonomy_node_ids": '["TEST-2026-1000"]',
                "description": "测试产出",
                "evidence_state": "supported",
                "evidence_ids": '["APP-E-001"]',
                "limitations": "",
            },
        )
        append_record(
            application,
            "应用节点",
            {
                "application_node_id": "APP-001",
                "name_zh": "合成应用",
                "name_en": "Synthetic application",
                "application_type": "manufacturing_input",
                "use_point_type": "equipment_control",
                "process_step": "test step",
                "output_product_ids": '["OUT-001"]',
                "description": "测试应用",
                "evidence_state": "supported",
                "evidence_ids": '["APP-E-001"]',
                "limitations": "",
            },
        )
        append_record(
            application,
            "需求原子",
            {
                "requirement_atom_id": "REQ-001",
                "application_node_id": "APP-001",
                "dimension": "function",
                "operator": "equals",
                "value": "synthetic control",
                "unit": "",
                "conditions": "test condition",
                "hardness": "hard",
                "evidence_state": "supported",
                "evidence_ids": '["APP-E-001"]',
                "limitations": "",
            },
        )
        registry_path = root / "00-管理" / "map-registry.json"
        registry = json.loads(registry_path.read_text(encoding="utf-8"))
        registry["shared_taxonomy"]["sha256"] = sha256(taxonomy)
        registry["shared_application_base"]["sha256"] = sha256(application)
        write_json(registry_path, registry)

    def populate_company_route(self, company_root: Path) -> None:
        company_map = company_root / "company-industry-application-map.xlsx"
        append_record(
            company_map,
            "产品能力",
            {
                "capability_id": "CAP-001",
                "company_id": "ACME-001",
                "product_scope": "Synthetic control material",
                "product_fact_ids": '["ACME-001-K-0001"]',
                "product_source_ids": '["ACME-001-S-0001"]',
                "dimension": "function",
                "operator": "equals",
                "value": "synthetic control",
                "unit": "",
                "conditions": "test condition",
                "known_limits": "[]",
                "evidence_state": "supported",
            },
        )
        append_record(
            company_map,
            "路线候选",
            {
                "route_candidate_id": "ACME-001-R-0001",
                "company_id": "ACME-001",
                "product_scope": "Synthetic control material",
                "application_node_id": "APP-001",
                "taxonomy_node_ids": '["TEST-2026-1000"]',
                "output_product_ids": '["OUT-001"]',
                "use_point_or_process": "equipment_control/test step",
                "target_enterprise_activity": "produces synthetic output",
                "product_fact_ids": '["ACME-001-K-0001"]',
                "product_source_ids": '["ACME-001-S-0001"]',
                "application_evidence_ids": '["APP-E-001"]',
                "application_source_groups": '["APP-SOURCE-GROUP-1"]',
                "evidence_state": "supported",
                "technical_match_state": "satisfied",
                "known_limit_conflict": "false",
                "research_disposition": "active",
                "map_route_status": "路线候选",
                "geography_hypotheses": "[]",
                "geography_evidence_ids": "[]",
                "unresolved_conditions": "[]",
                "derivation_trace": "CAP-001 -> REQ-001 -> APP-001",
            },
        )
        append_record(
            company_map,
            "匹配明细",
            {
                "match_id": "MATCH-001",
                "route_candidate_id": "ACME-001-R-0001",
                "requirement_atom_id": "REQ-001",
                "capability_id": "CAP-001",
                "match_state": "satisfied",
                "condition_compatibility": "satisfied",
                "process_interface_compatibility": "satisfied",
                "limit_conflict": "false",
                "product_fact_ids": '["ACME-001-K-0001"]',
                "application_evidence_ids": '["APP-E-001"]',
                "rationale": "Synthetic exact match",
            },
        )
        append_record(
            company_map,
            "覆盖台账",
            {
                "coverage_id": "COV-001",
                "company_id": "ACME-001",
                "product_scope": "Synthetic control material",
                "coverage_object_type": "capability",
                "coverage_object_id": "CAP-001",
                "coverage_state": "mapped",
                "disposition": "active",
                "route_candidate_ids": '["ACME-001-R-0001"]',
                "gap": "",
                "reviewed_at": "2026-07-29",
            },
        )

    def create_valid_workspace(self, parent: Path) -> tuple[Path, Path]:
        root = self.initialize_root(parent)
        self.populate_shared_base(root)
        library, packet = self.create_product_library(parent)
        company_root = self.initialize_company(root, library, packet)
        self.populate_company_route(company_root)
        return root, company_root

    def validate(self, root: Path, company_id: str = "ACME-001"):
        return run_script(
            VALIDATE_SCRIPT,
            str(root),
            "--company-id",
            company_id,
            "--format",
            "json",
        )

    def test_initializer_creates_empty_shared_root_and_refuses_overwrite(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = self.initialize_root(Path(tmp))
            self.assertTrue((root / "01-共享行业骨架" / "industry-taxonomy.xlsx").is_file())
            self.assertTrue(
                (root / "02-共享应用知识" / "industry-application-base.xlsx").is_file()
            )
            self.assertTrue((root / "04-公司地图").is_dir())

            second = run_script(
                INIT_SCRIPT,
                "--mode",
                "root",
                "--destination",
                str(root),
            )
            self.assertNotEqual(second.returncode, 0)
            self.assertIn("DESTINATION_EXISTS", second.stderr + second.stdout)

    def test_two_company_domains_share_only_the_neutral_base(self):
        with tempfile.TemporaryDirectory() as tmp:
            parent = Path(tmp)
            root = self.initialize_root(parent)
            self.populate_shared_base(root)
            material_library, material_packet = self.create_product_library_for(
                parent, "MATERIAL-001", "Synthetic effect material"
            )
            drive_library, drive_packet = self.create_product_library_for(
                parent, "DRIVE-001", "Synthetic motor drive"
            )
            material_root = self.initialize_company(
                root,
                material_library,
                material_packet,
                company_id="MATERIAL-001",
                product_scope="Synthetic effect material",
            )
            drive_root = self.initialize_company(
                root,
                drive_library,
                drive_packet,
                company_id="DRIVE-001",
                product_scope="Synthetic motor drive",
            )

            registry = json.loads(
                (root / "00-管理" / "map-registry.json").read_text(encoding="utf-8")
            )
            self.assertEqual(
                {item["company_id"] for item in registry["companies"]},
                {"MATERIAL-001", "DRIVE-001"},
            )
            material_input = load_workbook(
                material_root / "company-industry-application-map.xlsx", read_only=True
            )["公司与输入"]["A3"].value
            drive_input = load_workbook(
                drive_root / "company-industry-application-map.xlsx", read_only=True
            )["公司与输入"]["A3"].value
            self.assertEqual(material_input, "MATERIAL-001")
            self.assertEqual(drive_input, "DRIVE-001")
            self.assertNotEqual(material_root, drive_root)

    def test_company_initializer_freezes_inputs_and_stays_company_scoped(self):
        with tempfile.TemporaryDirectory() as tmp:
            parent = Path(tmp)
            root = self.initialize_root(parent)
            library, packet = self.create_product_library(parent)
            company_root = self.initialize_company(root, library, packet)
            registry = json.loads(
                (root / "00-管理" / "map-registry.json").read_text(encoding="utf-8")
            )
            company = registry["companies"][0]
            self.assertEqual(company["company_id"], "ACME-001")
            self.assertEqual(company["product_packet_sha256"], sha256(packet))
            self.assertNotIn("OTHER-CO", json.dumps(registry))
            self.assertTrue(
                (company_root / "company-industry-application-map.xlsx").is_file()
            )
            export_registry = json.loads(
                (company_root / "route-pool-export-registry.json").read_text(
                    encoding="utf-8"
                )
            )
            self.assertEqual(export_registry["company_id"], "ACME-001")
            self.assertEqual(export_registry["exports"], [])

    def test_valid_workspace_passes_and_exports_controlled_packet(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            result = self.validate(root)
            self.assertEqual(result.returncode, 0, result.stderr + result.stdout)
            report = json.loads(result.stdout)
            self.assertEqual(report["status"], "PASS")

            output = company_root / "company-route-pool-packet.json"
            export = run_script(
                EXPORT_SCRIPT,
                str(root),
                "--company-id",
                "ACME-001",
                "--output",
                str(output),
            )
            self.assertEqual(export.returncode, 0, export.stderr + export.stdout)
            packet = json.loads(output.read_text(encoding="utf-8"))[
                "company_route_pool_packet"
            ]
            self.assertEqual(packet["company_id"], "ACME-001")
            self.assertEqual(packet["target_skill"], "foreign-trade-customer-development")
            self.assertEqual(packet["route_candidates"][0]["route_candidate_id"], "ACME-001-R-0001")
            self.assertTrue(packet["export_id"].startswith("ACME-001-EXPORT-"))
            self.assertEqual(
                packet["producer_registry_reference"]["export_id"],
                packet["export_id"],
            )
            registry_path = company_root / "route-pool-export-registry.json"
            export_registry = json.loads(registry_path.read_text(encoding="utf-8"))
            export_record = export_registry["exports"][0]
            self.assertEqual(export_record["export_id"], packet["export_id"])
            self.assertEqual(export_record["packet_sha256"], sha256(output))
            self.assertEqual(export_record["state"], "current")
            self.assertEqual(export_record["validator_version"], "1.1")
            self.assertEqual(
                export_record["producer_snapshot"]["company_map_sha256"],
                sha256(company_root / "company-industry-application-map.xlsx"),
            )
            def contains_key(value, target_key):
                if isinstance(value, dict):
                    return target_key in value or any(
                        contains_key(item, target_key) for item in value.values()
                    )
                if isinstance(value, list):
                    return any(contains_key(item, target_key) for item in value)
                return False

            self.assertFalse(contains_key(packet, "direction_status"))

    def test_tampered_exported_route_packet_is_rejected_by_producer_registry(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            output = company_root / "company-route-pool-packet.json"
            export = run_script(
                EXPORT_SCRIPT,
                str(root),
                "--company-id",
                "ACME-001",
                "--output",
                str(output),
            )
            self.assertEqual(export.returncode, 0, export.stderr + export.stdout)
            payload = json.loads(output.read_text(encoding="utf-8"))
            payload["company_route_pool_packet"]["prohibited_inference"] = []
            write_json(output, payload)

            result = self.validate(root)

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("ROUTE_PACKET_HASH_MISMATCH", result.stdout)

    def test_current_export_is_rejected_after_source_company_map_changes(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            output = company_root / "company-route-pool-packet.json"
            export = run_script(
                EXPORT_SCRIPT,
                str(root),
                "--company-id",
                "ACME-001",
                "--output",
                str(output),
            )
            self.assertEqual(export.returncode, 0, export.stderr + export.stdout)
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "target_enterprise_activity",
                "changed after export",
            )

            result = self.validate(root)

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("ROUTE_EXPORT_SOURCE_MAP_STALE", result.stdout)

    def test_missing_export_registry_is_rejected_after_company_initialization(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            registry_path = company_root / "route-pool-export-registry.json"
            self.assertTrue(registry_path.is_file(), "initializer must create registry")
            registry_path.unlink()

            result = self.validate(root)

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("ROUTE_EXPORT_REGISTRY_MISSING", result.stdout)

    def test_cross_company_route_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "company_id",
                "OTHER-CO",
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("CROSS_COMPANY_ROUTE", result.stdout)

    def test_company_input_snapshot_tampering_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "公司与输入",
                "company_id",
                "OTHER-CO",
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("COMPANY_INPUT_SNAPSHOT_MISMATCH", result.stdout)

    def test_unresolved_product_fact_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "product_fact_ids",
                '["ACME-001-K-9999"]',
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("UNRESOLVED_PRODUCT_FACT", result.stdout)

    def test_route_candidate_requires_supported_evidence_and_satisfied_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "technical_match_state",
                "unknown",
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("ROUTE_CANDIDATE_NOT_TECHNICALLY_SATISFIED", result.stdout)

    def test_salesperson_owned_status_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "map_route_status",
                "已确认可扫描",
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("ROUTE_STATUS_EXCEEDS_AUTHORITY", result.stdout)

    def test_circular_product_and_application_source_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "application_source_groups",
                '["ACME-001-S-0001"]',
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("CIRCULAR_SOURCE_DEPENDENCY", result.stdout)

    def test_changed_product_packet_hash_requires_revalidation(self):
        with tempfile.TemporaryDirectory() as tmp:
            parent = Path(tmp)
            root, _ = self.create_valid_workspace(parent)
            packet = (
                parent
                / "ACME-001-product-library"
                / "04-开发交接"
                / "product-development-fact-packet.json"
            )
            payload = json.loads(packet.read_text(encoding="utf-8"))
            payload["product_development_fact_packet"]["generated_at"] = "2026-07-30"
            write_json(packet, payload)
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("PRODUCT_PACKET_HASH_MISMATCH", result.stdout)

    def test_geography_hypothesis_requires_traceable_evidence(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "路线候选",
                "geography_hypotheses",
                '["United States"]',
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("GEOGRAPHY_HYPOTHESIS_EVIDENCE_REQUIRED", result.stdout)

    def test_every_capability_requires_a_coverage_disposition(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "覆盖台账",
                "coverage_object_id",
                "CAP-OTHER",
            )
            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("CAPABILITY_COVERAGE_MISSING", result.stdout)

    def test_route_candidate_must_cover_every_hard_requirement(self):
        with tempfile.TemporaryDirectory() as tmp:
            root, company_root = self.create_valid_workspace(Path(tmp))
            application = root / "02-共享应用知识" / "industry-application-base.xlsx"
            append_record(
                application,
                "需求原子",
                {
                    "requirement_atom_id": "REQ-002",
                    "application_node_id": "APP-001",
                    "dimension": "environment",
                    "operator": "equals",
                    "value": "synthetic environment",
                    "unit": "",
                    "conditions": "test condition",
                    "hardness": "hard",
                    "evidence_state": "supported",
                    "evidence_ids": '["APP-E-001"]',
                    "limitations": "",
                },
            )
            registry_path = root / "00-管理" / "map-registry.json"
            registry = json.loads(registry_path.read_text(encoding="utf-8"))
            application_hash = sha256(application)
            registry["shared_application_base"]["sha256"] = application_hash
            registry["companies"][0]["application_base_sha256"] = application_hash
            write_json(registry_path, registry)
            change_first_record(
                company_root / "company-industry-application-map.xlsx",
                "公司与输入",
                "application_base_sha256",
                application_hash,
            )

            result = self.validate(root)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("ROUTE_HARD_REQUIREMENT_UNMATCHED", result.stdout)


if __name__ == "__main__":
    unittest.main()
